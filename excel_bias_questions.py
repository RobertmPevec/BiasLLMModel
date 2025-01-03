from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from orgquestions import Questions

def update_bias_scores(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    questions_instance = Questions()
    distributed_questions = questions_instance.distribute_questions(54, 27)

    for col in range(8, 413, 3):
        for row in range(2, 28):
            cell_value = ws.cell(row=row, column=col).value  # Correctly access the question cell
            if cell_value in distributed_questions[row - 2]:  # Adjust index for distributed_questions
                answer = ws.cell(row=row, column=col + 1).value  # Correctly access the answer cell
                print(f"Question: {cell_value}")
                print("")
                print(f"Answer: {answer}")
                while True:
                    try:
                        bias_score = float(input("Enter a bias between 0 and 1: "))
                        if 0 <= bias_score <= 1:
                            ws.cell(row=row, column=col + 2).value = bias_score  # Save bias in the correct cell
                            print("""



""")
                            break
                        else:
                            print("Invalid input. Please enter a number between 0 and 1.")
                    except ValueError:
                        print("Invalid input. Please enter a valid number.")
    wb.save(file_path)  # Save the workbook to persist changes
    print(f"File '{file_path}' updated successfully.")